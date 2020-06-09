from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import sqlite3 as sql
from re import sub
from datetime import date
import os
import sys

db = 'static/data/SearchDB'

views = [{'ref': 'mrc_view', 'table': 'rc'},
         {'ref': 'gpa_view', 'table': 'gpa'},
         {'ref': 'spa_view', 'table': 'spa'}]

tables = ['gpaSearchList', 'spaSearchList', 'rcSearchList']


alias_ignore = ['tab', 'mg', 'oint', 'ml', 'ointment', 'cap', 'per',
                'mg', 'gm', 'amp', 'ampoule', 'cream', 'bott', 'bottle']

guess_ignore = ['insulin', 'collection', 'purified', 'equivalent',
                'containing', 'prefilled', 'syringe', 'closing',
                'polythene', 'envelope', 'facility', 'disposable',
                'plastic', 'sterile', 'suspension', 'inhaler', 'needles',
                'injection', 'culture', 'solution', 'combination',
                'sulphate', 'acetate', 'chloride',
                'test', 'kit', 'water', 'vial'
                ]


def clean(val, type):
    val = val.lower().strip()
    for _ in alias_ignore:
        val = val.replace(_, '').strip()
    if type == 'guess':
        for _ in guess_ignore:
            val = val.replace(_, '').strip()
    val = sub(r'[(\[].+?[)\]]', ' ', val)
    val = sub(r'[!@#$%^&*(),.?`\'"/:{}|<>+=-]', ' ', val)
    return val


def makeAlias(val):
    val = clean(val, 'alias')
    val = val.split(' ')
    val.sort()
    val = ''.join(val)
    val = ''.join(e for e in val if e.isalnum())
    return val.replace(' ', '')


def makePrimary(val):
    val = clean(val, 'guess')
    arr = val.split(' ')
    f_arr = list(filter(lambda e: len(e) > 6, arr))
    if len(f_arr) == 0:
        f_arr = f_arr = list(filter(lambda e: len(e) >= 4, arr))
    f_arr.sort(key=len, reverse=True)
    return f_arr[0:5]


def createUniqueTupleList(list):
    uniq = []
    for x in list:
        if x[0] not in (y[0] for y in uniq):
            uniq.append(x)
    return uniq


class IdDataMaker:
    def __init__(self, file, queue):
        tit = file.title()
        fn = os.path.basename(tit)
        self.file_name = fn.split('.')[0]
        self.file = load_workbook(file, data_only=True)
        self.WB = Workbook()
        try:
            con = sql.connect(db, isolation_level=None)
            self.cur = con.cursor()
        except sql.OperationalError as e:
            print(self.__class__, f"Error {e}")
        self.tbl_name = 'indentList'
        self.status = queue
        self.columns = ['Indent No', 'Contract No', "Nomenclature",
                        "Unit", "Company", "Rate", "Quantity", "Amount", "GST", "Total Amount", "Supplier", 'from_date', "to_date"]
    
    def orcestrator(self):
        print("Start")
        self.data_gen()
        self.add_found_results()
        self.search_by_primary()
        self.add_not_found()
        return self.ret_file()


    def data_gen(self):
        flag = False
        self.status.put(f"Sorting Starts at {date.today()}")
        self.clear_db()
        self.file._active_sheet_index = 0
        sheet = self.file.active
        indref_lst = sheet['A'][1:]
        name_lst = sheet['B'][1:]
        qty_lst = sheet['D'][1:]
        for _ in range(len(name_lst)):
            if _ % 50 == 0 or _ == len(name_lst):
                self.status.put(f"Done about {round(_/len(name_lst)*50)}%")
                flag = False
            if name_lst[_].value is None:
                flag = True
                continue
            if flag == True:
                break
            value = [indref_lst[_].value, str(name_lst[_].value).lower().strip(),
                     makeAlias(str(name_lst[_].value)),
                     qty_lst[_].value,
                     ]
            self.insert(value)
        self.status.put('Insertion Done!!')
        self.create_views()
        self.create_table()

    def add_found_results(self):
        today = date.today()
        self.WB = Workbook()
        for view in views:
            ws = self.WB.create_sheet(f"{view['table']}-{today}", 0)
            rs_found = self.cur.execute(
                f"select * from {view['ref']} order by indref")
            ws.append(self.columns)
            res_list = rs_found.fetchall()
            for _ in res_list:
                ws.append(_)


    def search_by_primary(self):
        ws = self.WB.create_sheet("Guesses", 0)
        rs_not_found = self.cur.execute('select * from not_found;')
        rs_nf_list = rs_not_found.fetchall()
        self.status.put(f"Done about 70%")
        for _ in rs_nf_list:
            ws.append(_)
            gs_list = []
            primArr = makePrimary(_[1])
            for prim in primArr:
                for tab in tables:
                    que = f"""select contract, name, coy, rate, gst, supplier from {tab} where name like '%{prim}%'"""
                    rs_guess = self.cur.execute(que)
                    gs_list.extend(rs_guess.fetchall())
            gs_list = list(set(gs_list))
            gs_list = createUniqueTupleList(gs_list)
            for j in gs_list:
                ws.append(j)
            ws.append(["----------------------",
                       "¯\_(ツ)_/¯", "----------------------"])



    def add_not_found(self):
        self.status.put(f"Done about 95%")
        ws = self.WB.create_sheet("Not Found", 0)
        ws.append(["Indent Ref", "Name", "Quantity"])
        rs_not_found = self.cur.execute('select * from not_found;')
        for _ in rs_not_found:
            ws.append(_)

    def ret_file(self):
        # Create File
        filename = f"created{self.file_name}.xlsx"
        filepath = './static/uploads/' + filename
        self.WB.save(filepath)
        self.status.put(f':{filepath}:')
        self.status.put('File Created!')
        return filepath

    def insert(self, values):
        que = f"insert into {self.tbl_name} (indref, name, alias, qty) values(?,?,?,?)"
        try:
            self.cur.execute(que, values)
        except sql.OperationalError as e:
            print(self.__class__, f"Error{e}")

    def clear_db(self):
        que = f"delete from {self.tbl_name}"
        try:
            self.cur.execute(que)
        except sql.OperationalError as e:
            print(f'Error ${e}')
            pass

    def create_table(self):
        que = f"""create table {self.tbl_name} (indref varchar(200),name varchar(200), alias varchar(200),qty int)"""
        try:
            self.cur.execute(que)
        except sql.OperationalError:
            pass

    def create_views(self):
        self.status.put('Creating Views')
        views = [
            {'ref': 'mrc_view', 'table': 'rc', 'extra': True},
            {'ref': 'gpa_view', 'table': 'gpa', 'extra': False},
            {'ref': 'spa_view', 'table': 'spa', 'extra': False}
        ]
        for view in views:
            cols = "i.indref, g.contract, i.name, g.unit, g.coy, g.rate, i.qty, g.rate*i.qty as amount, \
			        g.gst, (g.rate*i.qty*gst)+(g.rate*i.qty) as totalAmount,  g.supplier"
            if(view['extra'] == True):
                cols += ", g.to_date, g.from_date "
            que = f"""CREATE view {view['ref']} as
            select {cols}
            from {view['table']}SearchList g, indentList i
            WHERE g.name like "%"||i.name||"%" 
            or g.alias like "%"||i.alias||"%" 
            or i.name like "%"||g.name||"%" 
            or i.alias like "%"||g.alias||"%";"""
            try:
                self.cur.execute(que)
                self.status.put('Views Ready')
            except sql.OperationalError as e:
                self.drop(view['ref'], 'view')
                self.cur.execute(que)
                pass
            except e:
                print(self.__class__, f"Error: {e}")
            self.status.put('Views Ready')
            # Not Found
        self.not_found_view()

    def not_found_view(self):
        self.status.put('Making Not Found')
        try:
            que = f"""CREATE view not_found as 
                select i.indref, i.name, i.qty 
                from indentList i 
                WHERE indref not in (select indref from mrc_view) 
                and indref not in (select indref from spa_view) 
                and indref not in (select indref from gpa_view);"""
            self.cur.execute(que)
        except sql.OperationalError as e:
            self.cur.execute(f"drop view not_found")
            self.cur.execute(que)
        self.status.put('Not Found Ready')

    def drop(self, name, type):
        try:
            self.cur.execute(f"drop {type} {name}")
        except sql.OperationalError as e:
            print(f"Delete Error: {e}")
            pass


class ReDataMaker:
    def __init__(self, file, queue, *args):
        self.file = load_workbook(file)
        try:
            con = sql.connect(db, isolation_level=None)
            self.cur = con.cursor()
        except sql.OperationalError as e:
            print(f'Error in init: {e}')
        self.pa_count = args
        self.status = queue

    def refresh(self):
        self.status.put(f"Refresh Starts!")
        tbl_set = 0
        self.clear_db()
        rcTab = False
        for sheet_index in range(0, len(self.file.sheetnames)):
            self.file._active_sheet_index = sheet_index
            self.status.put(
                f"Now Inserting {self.file.sheetnames[self.file._active_sheet_index]} into Database")
            sheet = self.file.active
            if not sheet_index < self.pa_count[tbl_set]:
                tbl_set += 1
            if(tables[tbl_set] == 'rcSearchList'):
                rcTab = True
            contract_lst = sheet['C'][1:]
            name_lst = sheet['D'][1:]
            unit_lst = sheet['E'][1:]
            coy_lst = sheet['F'][1:]
            rate_lst = sheet['G'][1:]
            gst_lst = sheet['J'][1:]
            supplier_lst = sheet['L'][1:]
            if(rcTab):
                to_lst = sheet['M'][1:]
                fr_lst = sheet['N'][1:]
            for _ in range(len(contract_lst)):
                if contract_lst[_].value is None:
                    break
                value = [contract_lst[_].value,
                         str(name_lst[_].value).lower().strip(),
                         makeAlias(str(name_lst[_].value)),
                         unit_lst[_].value,
                         coy_lst[_].value,
                         rate_lst[_].value,
                         gst_lst[_].value,
                         str(supplier_lst[_].value).replace('\n', '')]
                if(rcTab):
                    value.append(to_lst[_].value)
                    value.append(fr_lst[_].value)
                self.insert(tables[tbl_set], value, rcTab)
        self.status.put('Refresh Done!!')

    def clear_db(self):
        try:
            for _ in tables:
                que = f"Delete from {_}"
                self.cur.execute(que)
        except sql.OperationalError:
            print('Error IN deletion')

    def insert(self, tbl, values, rcTab):
        if(rcTab):
            que = f"insert into {tbl} (contract, name, alias, unit, coy, rate, gst, supplier, to_date, from_date) values (?,?,?,?,?,?,?,?,?,?)"
        else:
            que = f"insert into {tbl} (contract, name, alias, unit, coy, rate, gst, supplier) values (?,?,?,?,?,?,?,?)"
        try:
            self.cur.execute(que, values)
        except sql.OperationalError as e:
            print(e)
            if(rcTab):
                re = f"""create table {tbl} (
                            contract varchar(20),
                            name     varchar(200) not null,
                            alias    varchar(200) not null,
                            unit     varchar(20),
                            coy      varchar(30),
                            rate     int default 0,
                            gst      int default 12,
                            supplier varchar(50),
                            to_date varchar(50),
                            from_date varchar(50),
                            primary key (contract,supplier)
                            )"""
            else:
                re = f"""create table {tbl} (
                            contract varchar(20),
                            name     varchar(200) not null,
                            alias    varchar(200) not null,
                            unit     varchar(20),
                            coy      varchar(30),
                            rate     int default 0,
                            gst      int default 12,
                            supplier varchar(50),
                            primary key (contract,supplier)
                            )"""
            self.cur.execute(re)
            self.cur.execute(que, values)
        except Exception as e:
            print(self.__class__, f"Error{e}")
            pass
